from flask import Flask, jsonify, request, send_file
import joblib
import pandas as pd
import numpy as np
import os
import networkx as nx
import pulp
from shapely import wkt
from datetime import datetime
import io
import uuid
import json
import threading
import math
import folium
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch# Import local modules
import sys
base_dir = os.path.dirname(os.path.abspath(__file__))
sys.path.append(os.path.join(base_dir, "scratch"))
sys.path.append(os.path.join(base_dir, "services"))
from gis_twin import G, JUNCTIONS, CORRIDOR_EDGES, calculate_route, calculate_k_shortest_paths, simulate_flow_diversion, get_corridies_with_geometries
from optimizer import optimize_resources
from btp_event_service import get_btp_events, save_btp_import

app = Flask(__name__)

# Paths
model_dir = os.path.join(base_dir, "models")
data_dir = os.path.join(base_dir, "data")
os.makedirs(data_dir, exist_ok=True)
csv_path = os.path.join(base_dir, "Astram event data_anonymized - Astram event data_anonymizedb40ac87.csv")
feedback_csv_path = os.path.join(base_dir, "feedback_data.csv")
feedback_xlsx_path = os.path.join(data_dir, "post_event_feedback.xlsx")
btp_xlsx_path = os.path.join(data_dir, "imported_btp_events.xlsx")
audit_xlsx_path = os.path.join(data_dir, "audit_log.xlsx")
registry_path = os.path.join(data_dir, "model_versions.json")
active_txt_path = os.path.join(model_dir, "active_version.txt")

playbooks_dir = os.path.join(base_dir, "playbooks")
os.makedirs(playbooks_dir, exist_ok=True)

@app.route('/playbooks/<path:filename>')
def serve_playbook(filename):
    return send_file(os.path.join(playbooks_dir, filename))

# Global state for loaded models and encoders
models = {}
encoders = {}
active_incidents = []  # In-memory store for active incidents
historical_stats = {}
retraining_status = {}  # {job_id: {status, progress, message, result}}
audit_log_memory = []  # In-memory audit log buffer
impact_explainer = None  # SHAP explainer for impact model

# ── Excel helpers ──────────────────────────────────────────────────────────────
def _feedback_xlsx_write(row_data: dict):
    """Append a feedback row to the Excel file."""
    import openpyxl
    COLUMNS = [
        'feedback_id','incident_id','event_cause','event_type','junction','corridor',
        'predicted_impact','actual_impact','predicted_duration','actual_duration',
        'predicted_resource_count','actual_resource_count',
        'road_closure_predicted','road_closure_actual',
        'feedback_timestamp','submitted_by','approval_status','approved_by','approval_timestamp'
    ]
    if os.path.exists(feedback_xlsx_path):
        wb = openpyxl.load_workbook(feedback_xlsx_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append(COLUMNS)
    ws.append([row_data.get(c, '') for c in COLUMNS])
    wb.save(feedback_xlsx_path)

def _feedback_xlsx_read() -> list:
    import openpyxl
    if not os.path.exists(feedback_xlsx_path):
        return []
    wb = openpyxl.load_workbook(feedback_xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]

def _feedback_xlsx_update(feedback_id: str, update: dict):
    import openpyxl
    if not os.path.exists(feedback_xlsx_path):
        return False
    wb = openpyxl.load_workbook(feedback_xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index('feedback_id') + 1
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col - 1].value) == feedback_id:
            for key, val in update.items():
                if key in headers:
                    row[headers.index(key)].value = val
            break
    wb.save(feedback_xlsx_path)
    return True

def _audit_log_xlsx_write(entry: dict):
    import openpyxl
    COLS = ['log_id','timestamp','user_role','action','module','details']
    if os.path.exists(audit_xlsx_path):
        wb = openpyxl.load_workbook(audit_xlsx_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AuditLog"
        ws.append(COLS)
    ws.append([entry.get(c, '') for c in COLS])
    wb.save(audit_xlsx_path)

def _audit_log_xlsx_read() -> list:
    import openpyxl
    if not os.path.exists(audit_xlsx_path):
        return []
    wb = openpyxl.load_workbook(audit_xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]

def log_audit(user_role: str, action: str, module: str, details: dict = None):
    """Write audit entry to memory + xlsx."""
    entry = {
        'log_id': str(uuid.uuid4())[:8],
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user_role': user_role,
        'action': action,
        'module': module,
        'details': json.dumps(details or {})
    }
    audit_log_memory.append(entry)
    try:
        _audit_log_xlsx_write(entry)
    except Exception:
        pass

def _load_registry() -> list:
    if os.path.exists(registry_path):
        with open(registry_path, 'r') as f:
            return json.load(f)
    return []

def load_intelligence_components():
    global models, encoders, cause_means, historical_stats, impact_explainer
    print("Loading serialized ML components...")
    try:
        models['impact'] = joblib.load(os.path.join(model_dir, 'best_impact_model.joblib'))
        models['duration'] = joblib.load(os.path.join(model_dir, 'best_duration_model.joblib'))
        models['closure'] = joblib.load(os.path.join(model_dir, 'best_closure_model.joblib'))
        encoders = joblib.load(os.path.join(model_dir, 'label_encoders.joblib'))
        cause_means = joblib.load(os.path.join(model_dir, 'cause_means.joblib'))
        
        import shap
        impact_explainer = shap.TreeExplainer(models['impact'])
        
        print("All ML models and SHAP explainers loaded successfully.")
    except Exception as e:
        print(f"Error loading ML components: {e}")

    # Load high-level stats from CSV
    try:
        df = pd.read_csv(csv_path)
        requires_closure = (df['requires_road_closure'].astype(str).str.lower().str.strip() == 'true').sum()
        historical_stats = {
            'total_events': len(df),
            'road_closures': int(requires_closure),
            'planned_events': int((df['event_type'] == 'planned').sum()),
            'unplanned_events': int((df['event_type'] == 'unplanned').sum()),
        }
        print("Historical stats pre-loaded.")
    except Exception as e:
        print(f"Error loading historical CSV stats: {e}")

# Pre-load components on startup
load_intelligence_components()

# ── Load Global Graph on Startup ──
global_G = None
try:
    print("Loading OSM graphml...")
    import osmnx as ox
    global_G = ox.load_graphml(os.path.join(data_dir, "bengaluru_graph.graphml"))
    print("OSM Graphml loaded successfully.")
except Exception as e:
    print(f"Could not load OSM graph: {e}")

# ── API Endpoints ──────────────────────────────────────────────────────────────

@app.route('/api/dashboard_stats', methods=['GET'])
def get_dashboard_stats():
    return jsonify({
        'success': True,
        'stats': historical_stats,
        'active_count': len(active_incidents)
    })

@app.route('/api/model-versions', methods=['GET'])
def get_model_versions():
    reg = _load_registry()
    return jsonify({"success": True, "versions": reg})

@app.route('/api/active-model', methods=['GET'])
def get_active_model():
    reg = _load_registry()
    active = next((v for v in reg if v.get("is_active")), None)
    if active:
        return jsonify({"success": True, "active": active})
    return jsonify({"success": False, "message": "No active model found"})

@app.route('/api/rollback-model', methods=['POST'])
def rollback_model():
    data = request.json or {}
    target_version = data.get("version")
    if not target_version:
        return jsonify({"success": False, "message": "Version required"}), 400
        
    import shutil
    imp_src = os.path.join(model_dir, f'xgb_imp_{target_version}.pkl')
    dur_src = os.path.join(model_dir, f'xgb_dur_{target_version}.pkl')
    cls_src = os.path.join(model_dir, f'lgb_cls_{target_version}.pkl')
    
    if not (os.path.exists(imp_src) and os.path.exists(dur_src) and os.path.exists(cls_src)):
        return jsonify({"success": False, "message": f"Physical model files for {target_version} not found"}), 404
        
    try:
        shutil.copy(imp_src, os.path.join(model_dir, 'best_impact_model.joblib'))
        shutil.copy(dur_src, os.path.join(model_dir, 'best_duration_model.joblib'))
        shutil.copy(cls_src, os.path.join(model_dir, 'best_closure_model.joblib'))
        
        with open(active_txt_path, "w") as f:
            f.write(target_version)
            
        reg = _load_registry()
        for v in reg:
            v["is_active"] = (v["version"] == target_version)
        with open(registry_path, "w") as f:
            json.dump(reg, f, indent=2)
            
        load_intelligence_components()
        log_audit("Admin", "Model Rollback", "ML Engine", {"version": target_version})
        
        return jsonify({"success": True, "message": f"Successfully rolled back to {target_version}"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

@app.route('/api/predict', methods=['POST'])
def predict_incident():
    """
    Inputs: event_cause, corridor, zone, junction, hour, dow, event_type, priority
    """
    data = request.json or {}
    
    # Extract inputs
    cause = data.get('event_cause', 'vehicle_breakdown')
    corr = data.get('corridor', 'Unknown')
    zone = data.get('zone', 'Unknown')
    junc = data.get('junction', 'Unknown')
    hour = int(data.get('hour', 12))
    dow = int(data.get('dow', 0))
    etype = data.get('event_type', 'unplanned')
    priority = data.get('priority', 'Low')
    
    # Calculate historical severity
    hist_sev = cause_means.get(cause, 3.5)
    
    # Prepare input DataFrame for encoders & models
    input_data = pd.DataFrame([{
        'event_cause': cause,
        'corridor': corr,
        'zone': zone,
        'junction': junc,
        'hour': hour,
        'dow': dow,
        'event_type': etype,
        'priority': priority,
        'historical_severity': hist_sev
    }])
    
    # Encode variables
    try:
        for col, le in encoders.items():
            val = input_data[col].iloc[0]
            if val not in le.classes_:
                # Use first class or default if unseen in training
                input_data[col] = le.transform([le.classes_[0]])
            else:
                input_data[col] = le.transform([val])
    except Exception as e:
        return jsonify({'success': False, 'error': f"Encoding failed: {e}"})
        
    # Run Predictions
    try:
        # Predict Impact Score
        impact_score = float(models['impact'].predict(input_data)[0])
        # Predict Duration (trained on log scale)
        log_dur = float(models['duration'].predict(input_data)[0])
        expected_dur = float(np.expm1(log_dur))
        # Predict Closure Probability
        closure_prob = float(models['closure'].predict_proba(input_data)[0][1])
        
        # Calculate SHAP for impact model
        global impact_explainer
        shap_vals = impact_explainer.shap_values(input_data)[0]
        feature_names = input_data.columns.tolist()
        top_drivers_raw = sorted(zip(feature_names, shap_vals.tolist()), key=lambda x: abs(x[1]), reverse=True)[:3]
        
        shap_drivers = []
        for feat, val in top_drivers_raw:
            direction = "up" if val > 0 else "down"
            display_feat = feat.replace('_', ' ').title()
            shap_drivers.append({"feature": display_feat, "value": abs(val), "direction": direction})
        
        # Round values for display
        impact_score = min(10.0, max(1.0, round(impact_score, 1)))
        expected_dur = max(10, int(round(expected_dur)))
        closure_prob = round(closure_prob, 2)
        
        return jsonify({
            'success': True,
            'prediction': {
                'impact_score': impact_score,
                'closure_probability': closure_prob,
                'expected_duration_min': expected_dur,
                'historical_severity': round(hist_sev, 2),
                'shap_drivers': shap_drivers
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': f"Prediction model execution failed: {e}"})

@app.route('/api/corridors', methods=['GET'])
def get_corridors_api():
    try:
        corridors = get_corridies_with_geometries()
        return jsonify({
            'success': True,
            'corridors': corridors
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/api/route', methods=['POST'])
def calculate_network_route():
    data = request.json or {}
    source = data.get('source_junction')
    target = data.get('target_junction')
    closed_edges_raw = data.get('closed_edges', []) # list of lists, e.g. [['nodeA', 'nodeB']]
    
    closed_edges = [tuple(edge) for edge in closed_edges_raw]
    
    if not source or not target:
        return jsonify({'success': False, 'error': "Source and target junctions are required."})
        
    if source not in JUNCTIONS or target not in JUNCTIONS:
        return jsonify({'success': False, 'error': "Invalid source or target junction name."})
        
    # Standard Shortest Route
    route_res = calculate_route(source, target, closed_edges)
    
    if not route_res['success']:
        return jsonify(route_res)
        
    # K-Shortest Paths (alternatives)
    paths_res = calculate_k_shortest_paths(source, target, k=3, closed_edges=closed_edges)
    
    # If there is a closure, calculate flow diversion and secondary congestion
    flow_diversions = []
    if closed_edges:
        for edge in closed_edges:
            # Estimate diversion paths and loads
            divs = simulate_flow_diversion(edge, base_incident_impact=7.0)
            flow_diversions.extend(divs)
            
    return jsonify({
        'success': True,
        'primary_route': {
            'path': route_res['path'],
            'distance_km': round(route_res['distance'], 2),
            'time_min': round(route_res['time_min'], 1),
            'coordinates': route_res['coordinates']
        },
        'alternatives': paths_res.get('paths', []),
        'flow_diversions': flow_diversions
    })

# ── Feature: NetworkX Cascade Simulation ──────────────────────────────────────
global_G = None

def get_graph():
    global global_G
    if global_G is None:
        graph_path = os.path.join(data_dir, "bengaluru_graph.graphml")
        if os.path.exists(graph_path):
            global_G = nx.read_graphml(graph_path)
        else:
            global_G = nx.MultiDiGraph()
    return global_G

def nearest_node(G, lat, lon):
    min_dist = float('inf')
    best_node = None
    for n, data in G.nodes(data=True):
        if 'y' in data and 'x' in data:
            dist = (float(data['y']) - lat)**2 + (float(data['x']) - lon)**2
            if dist < min_dist:
                min_dist = dist
                best_node = n
    return best_node

@app.route('/api/simulate-diversion', methods=['POST'])
def simulate_diversion():
    data = request.json or {}
    blocked_corridor = data.get('blocked_corridor', '')
    origin_name = data.get('origin_junction')
    dest_name = data.get('dest_junction')
    k = int(data.get('k', 3))

    if origin_name not in JUNCTIONS or dest_name not in JUNCTIONS:
        return jsonify({'success': False, 'error': 'Invalid source or destination junctions'})

    G = get_graph()
    if len(G.nodes) == 0:
        return jsonify({'success': False, 'error': 'Graph data not found'})

    origin_node = nearest_node(G, float(JUNCTIONS[origin_name]['lat']), float(JUNCTIONS[origin_name]['lon']))
    dest_node = nearest_node(G, float(JUNCTIONS[dest_name]['lat']), float(JUNCTIONS[dest_name]['lon']))

    # Remove blocked corridor
    H = G.copy()
    blocked_edges = []
    blocked_geojson_coords = []
    for u, v, key, edge_data in list(H.edges(keys=True, data=True)):
        if blocked_corridor and blocked_corridor.lower() in str(edge_data.get('name', '')).lower():
            blocked_edges.append((u, v, key))
            if 'geometry' in edge_data:
                geom = wkt.loads(edge_data['geometry'])
                blocked_geojson_coords.extend([list(c) for c in geom.coords])
            else:
                n1 = H.nodes[u]
                n2 = H.nodes[v]
                if 'x' in n1 and 'y' in n1:
                    blocked_geojson_coords.append([float(n1['x']), float(n1['y'])])
                if 'x' in n2 and 'y' in n2:
                    blocked_geojson_coords.append([float(n2['x']), float(n2['y'])])
                    
    H.remove_edges_from(blocked_edges)

    # Calculate k-shortest
    try:
        from itertools import islice
        if isinstance(H, nx.MultiDiGraph):
            H_di = nx.DiGraph(H)
        else:
            H_di = H
            
        paths = list(islice(nx.shortest_simple_paths(H_di, origin_node, dest_node, weight='length'), k))
    except nx.NetworkXNoPath:
        return jsonify({'success': False, 'error': 'No alternate path exists after removing the corridor.'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

    # Score paths
    alternatives = []
    for i, path_nodes in enumerate(paths):
        coords = []
        corridors_passed = set()
        dist_m = 0.0
        
        for idx in range(len(path_nodes)-1):
            u, v = path_nodes[idx], path_nodes[idx+1]
            if H.has_edge(u, v):
                if isinstance(H, nx.MultiDiGraph):
                    edge_data = H[u][v][list(H[u][v].keys())[0]]
                else:
                    edge_data = H[u][v]
                    
                dist_m += float(edge_data.get('length', 0))
                c_name = str(edge_data.get('name', 'Unnamed'))
                if c_name != 'Unnamed':
                    corridors_passed.add(c_name)
                    
                if 'geometry' in edge_data:
                    geom = wkt.loads(edge_data['geometry'])
                    coords.extend([list(c) for c in geom.coords])
                else:
                    n1 = H.nodes[u]
                    if 'x' in n1 and 'y' in n1:
                        coords.append([float(n1['x']), float(n1['y'])])
        
        # Simulate secondary load percentage (historical context approx)
        load_pct = min(99.0, (len(corridors_passed) * 8.4 + (i * 22.5) + (dist_m / 1000.0)))
        rec = "PREFERRED" if load_pct < 25 else ("ACCEPTABLE" if load_pct <= 40 else "AVOID")
        
        alternatives.append({
            'rank': i+1,
            'distance_m': round(dist_m, 2),
            'geojson': {'type': 'LineString', 'coordinates': coords},
            'passes_through': list(corridors_passed)[:3],
            'secondary_load_pct': round(load_pct, 1),
            'recommendation': rec
        })
        
    b_coords = []
    for c in blocked_geojson_coords:
        if not b_coords or c != b_coords[-1]:
            b_coords.append(c)

    return jsonify({
        'success': True,
        'blocked_segment': {
            'corridor': blocked_corridor,
            'geojson': {'type': 'LineString', 'coordinates': b_coords}
        },
        'alternate_routes': alternatives
    })

@app.route('/api/optimize', methods=['POST'])
def run_resource_optimization():
    data = request.json or {}
    event_cause = data.get('event_cause', '')
    corridor = data.get('corridor', '')
    impact_score = float(data.get('impact_score', 0))
    closure_probability = float(data.get('closure_probability', 0))
    available_officers = int(data.get('available_officers', 20))
    available_barricades = int(data.get('available_barricades', 100))

    # Identify affected junctions based on the corridor
    affected_junc_names = set()
    for u, v, corr_name in CORRIDOR_EDGES:
        if corr_name == corridor:
            affected_junc_names.add(u)
            affected_junc_names.add(v)
            
    # Load junction metadata
    key_junc_path = os.path.join(data_dir, "key_junctions.json")
    junc_meta = {}
    if os.path.exists(key_junc_path):
        try:
            with open(key_junc_path, 'r') as f:
                for j in json.load(f):
                    junc_meta[j['name']] = j
        except Exception:
            pass
            
    affected_junctions = []
    for name in affected_junc_names:
        if name in junc_meta:
            affected_junctions.append(junc_meta[name])
        elif name in JUNCTIONS:
            # Fallback
            affected_junctions.append({
                "name": name,
                "lat": JUNCTIONS[name][0],
                "lon": JUNCTIONS[name][1],
                "degree": 3 # Default degree
            })
            
    if not affected_junctions:
        return jsonify({
            'status': 'infeasible',
            'total_officers': 0,
            'total_barricades': 0,
            'junction_deployments': [],
            'objective_value': 0,
            'coverage_warning': f"No junctions found for corridor {corridor}"
        })

    prob = pulp.LpProblem("Live_Optimization", pulp.LpMinimize)
    
    officers = pulp.LpVariable.dicts("Officers", [j['name'] for j in affected_junctions], lowBound=0, cat='Integer')
    barricades = pulp.LpVariable.dicts("Barricades", [j['name'] for j in affected_junctions], lowBound=0, cat='Integer')
    
    # Objective
    prob += pulp.lpSum([officers[j['name']] for j in affected_junctions]) + 0.5 * pulp.lpSum([barricades[j['name']] for j in affected_junctions])
    
    # Constraints
    prob += pulp.lpSum([officers[j['name']] for j in affected_junctions]) <= available_officers
    prob += pulp.lpSum([barricades[j['name']] for j in affected_junctions]) <= available_barricades
    
    min_off_total = 0
    for j in affected_junctions:
        j_name = j['name']
        j_load = j.get('degree', 3)
        # min_coverage constraint based on load and impact
        min_cov = max(1, math.ceil(j_load * impact_score * 0.1))
        
        prob += officers[j_name] >= min_cov
        min_off_total += min_cov
        
        # barricade constraint
        if closure_probability > 0.6:
            prob += barricades[j_name] >= 2
        elif closure_probability > 0.3:
            prob += barricades[j_name] >= 1
        else:
            prob += barricades[j_name] >= 0
            
        if impact_score >= 7:
            prob += officers[j_name] >= 2
            
    # Solve
    prob.solve(pulp.PULP_CBC_CMD(msg=False))
    
    status = pulp.LpStatus[prob.status]
    
    if status != 'Optimal':
        # If infeasible, report warning
        return jsonify({
            'status': 'infeasible',
            'total_officers': 0,
            'total_barricades': 0,
            'junction_deployments': [],
            'objective_value': 0,
            'coverage_warning': f"Insufficient officers — minimum {min_off_total} required"
        })
        
    deployments = []
    tot_off = 0
    tot_bar = 0
    for j in affected_junctions:
        j_name = j['name']
        off = int(officers[j_name].varValue)
        bar = int(barricades[j_name].varValue)
        tot_off += off
        tot_bar += bar
        
        priority = "LOW"
        if off >= 4:
            priority = "HIGH"
        elif off >= 2:
            priority = "MEDIUM"
            
        deployments.append({
            "junction_name": j_name,
            "lat": j['lat'],
            "lon": j['lon'],
            "officers": off,
            "barricades": bar,
            "priority": priority
        })
        
    return jsonify({
        'status': 'optimal',
        'total_officers': tot_off,
        'total_barricades': tot_bar,
        'junction_deployments': deployments,
        'objective_value': pulp.value(prob.objective),
        'coverage_warning': None
    })

@app.route('/api/incidents', methods=['GET', 'POST'])
def handle_incidents():
    global active_incidents
    if request.method == 'GET':
        return jsonify({
            'success': True,
            'incidents': active_incidents
        })
    else:
        # Create a new active incident
        data = request.json or {}
        cause = data.get('event_cause', 'vehicle_breakdown')
        corr = data.get('corridor', 'Unknown')
        zone = data.get('zone', 'Unknown')
        junc = data.get('junction', 'Unknown')
        etype = data.get('event_type', 'unplanned')
        priority = data.get('priority', 'Low')
        
        now = datetime.now()
        hour = now.hour
        dow = now.weekday()
        
        # Calculate predicted impact
        hist_sev = cause_means.get(cause, 3.5)
        input_data = pd.DataFrame([{
            'event_cause': cause, 'corridor': corr, 'zone': zone, 'junction': junc,
            'hour': hour, 'dow': dow, 'event_type': etype, 'priority': priority,
            'historical_severity': hist_sev
        }])
        
        try:
            for col, le in encoders.items():
                val = input_data[col].iloc[0]
                if val not in le.classes_:
                    input_data[col] = le.transform([le.classes_[0]])
                else:
                    input_data[col] = le.transform([val])
            
            imp_score = float(models['impact'].predict(input_data)[0])
            log_dur = float(models['duration'].predict(input_data)[0])
            expected_dur = float(np.expm1(log_dur))
            closure_prob = float(models['closure'].predict_proba(input_data)[0][1])
            
            imp_score = min(10.0, max(1.0, round(imp_score, 1)))
            expected_dur = max(10, int(round(expected_dur)))
            closure_prob = round(closure_prob, 2)
            
            global impact_explainer
            shap_vals = impact_explainer.shap_values(input_data)[0]
            feature_names = input_data.columns.tolist()
            top_drivers_raw = sorted(zip(feature_names, shap_vals.tolist()), key=lambda x: abs(x[1]), reverse=True)[:3]
            
            shap_drivers = []
            for feat, val in top_drivers_raw:
                direction = "up" if val > 0 else "down"
                display_feat = feat.replace('_', ' ').title()
                shap_drivers.append({"feature": display_feat, "value": abs(val), "direction": direction})
                
        except Exception as e:
            return jsonify({'success': False, 'error': f"Failed to predict new incident: {e}"})
            
        incident = {
            'id': f"INC{len(active_incidents) + 1:04d}",
            'event_cause': cause,
            'corridor': corr,
            'zone': zone,
            'junction': junc,
            'event_type': etype,
            'priority': priority,
            'start_datetime': now.strftime("%Y-%m-%d %H:%M:%S"),
            'hour': hour,
            'dow': dow,
            'impact_score': imp_score,
            'closure_probability': closure_prob,
            'expected_duration_min': expected_dur,
            'status': 'active',
            'shap_drivers': shap_drivers
        }
        
        active_incidents.append(incident)
        return jsonify({
            'success': True,
            'incident': incident
        })

@app.route('/api/road-network', methods=['GET'])
def get_road_network():
    path = os.path.join(data_dir, "road_network.geojson")
    if os.path.exists(path):
        return send_file(path, mimetype='application/json')
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/junctions', methods=['GET'])
def get_junctions():
    path = os.path.join(data_dir, "key_junctions.json")
    if os.path.exists(path):
        return send_file(path, mimetype='application/json')
    return jsonify({'error': 'Not found'}), 404

@app.route('/api/incidents/<inc_id>', methods=['DELETE'])
def resolve_incident(inc_id):
    global active_incidents
    incident = None
    for inc in active_incidents:
        if inc['id'] == inc_id:
            incident = inc
            break
            
    if not incident:
        return jsonify({'success': False, 'error': "Incident not found."}), 404
        
    incident['status'] = 'resolved'
    incident['resolved_datetime'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # Remove from active lists
    active_incidents = [i for i in active_incidents if i['id'] != inc_id]
    
    return jsonify({
        'success': True,
        'message': f"Incident {inc_id} resolved.",
        'resolved_incident': incident
    })

@app.route('/api/conflicts', methods=['GET'])
def check_active_conflicts():
    warnings = []
    
    # 1. Time & Corridor overlap conflicts
    corridor_active = {}
    for inc in active_incidents:
        corr = inc['corridor']
        if corr != 'Unknown' and corr != 'Non-corridor':
            if corr in corridor_active:
                corridor_active[corr].append(inc['id'])
            else:
                corridor_active[corr] = [inc['id']]
                
    for corr, ids in corridor_active.items():
        if len(ids) > 1:
            warnings.append(f"⚠️ Corridor conflict: Multiple active incidents {', '.join(ids)} on '{corr}' corridor.")
            
    # 2. Adjacent junction bottlenecks
    for i in range(len(active_incidents)):
        for j in range(i+1, len(active_incidents)):
            node1 = active_incidents[i]['junction']
            node2 = active_incidents[j]['junction']
            if node1 != 'Unknown' and node2 != 'Unknown':
                if G.has_edge(node1, node2) or G.has_edge(node2, node1):
                    warnings.append(f"⚠️ Junction bottleneck: Active incidents {active_incidents[i]['id']} ({node1}) and {active_incidents[j]['id']} ({node2}) are directly adjacent.")
                    
    # 3. Resource capacity shortfalls
    if active_incidents:
        res_opt = optimize_resources(active_incidents, 30, 50)
        target_officers = sum(round(inc['impact_score'] * 1.5) for inc in active_incidents)
        if target_officers > 30:
            warnings.append(f"🚨 Resource deficit: Required {target_officers} officers to cover all active safety margins, but only 30 are available.")
            
    return jsonify({
        'success': True,
        'warnings': warnings
    })

@app.route('/api/feedback', methods=['POST'])
def save_feedback():
    data = request.json or {}
    inc_id = data.get('incident_id')
    actual_imp = float(data.get('actual_impact', 5.0))
    actual_dur = float(data.get('actual_duration', 60.0))
    actual_close = str(data.get('actual_closure', 'False')).lower() == 'true'
    
    # Write to feedback log file
    log_header = not os.path.exists(feedback_csv_path)
    try:
        feedback_df = pd.DataFrame([{
            'incident_id': inc_id,
            'actual_impact': actual_imp,
            'actual_duration': actual_dur,
            'actual_closure': actual_close,
            'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }])
        feedback_df.to_csv(feedback_csv_path, mode='a', header=log_header, index=False)
        return jsonify({'success': True, 'message': "Post-event learning data logged successfully."})
    except Exception as e:
        return jsonify({'success': False, 'error': f"Failed to log feedback: {e}"})

@app.route('/api/playbook/<incident_id>', methods=['GET'])
def get_pdf_playbook(incident_id):
    """
    Generates a government-grade printable PDF playbook.
    """
    # Find incident
    incident = None
    for inc in active_incidents:
        if inc['id'] == incident_id:
            incident = inc
            break
            
    if not incident:
        # Mock up one if requesting an inactive ID for demonstration
        incident = {
            'id': incident_id,
            'event_cause': 'public_event',
            'corridor': 'CBD 2',
            'zone': 'Central Zone 2',
            'junction': 'MekhriCircle',
            'event_type': 'planned',
            'priority': 'High',
            'start_datetime': '2026-06-16 19:00:00',
            'impact_score': 8.2,
            'closure_probability': 0.85,
            'expected_duration_min': 240
        }
        
    from reportlab.lib.pagesizes import letter
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib import colors
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    story = []
    
    styles = getSampleStyleSheet()
    
    title_style = ParagraphStyle(
        'GovTitle',
        parent=styles['Heading1'],
        fontName='Helvetica-Bold',
        fontSize=20,
        leading=22,
        textColor=colors.HexColor('#0f172a'),
        spaceAfter=15
    )
    
    h2_style = ParagraphStyle(
        'GovH2',
        parent=styles['Heading2'],
        fontName='Helvetica-Bold',
        fontSize=13,
        leading=16,
        textColor=colors.HexColor('#1e293b'),
        spaceBefore=12,
        spaceAfter=6
    )
    
    body_style = ParagraphStyle(
        'GovBody',
        parent=styles['Normal'],
        fontName='Helvetica',
        fontSize=10,
        leading=14,
        textColor=colors.HexColor('#334155')
    )
    
    bold_style = ParagraphStyle(
        'GovBold',
        parent=body_style,
        fontName='Helvetica-Bold'
    )
    
    # Document Header
    story.append(Paragraph("BENGALURU TRAFFIC POLICE CONGESTION SYSTEM", ParagraphStyle('Sub', fontName='Helvetica', fontSize=9, leading=10, textColor=colors.HexColor('#475569'))))
    story.append(Paragraph(f"OPERATIONAL PLAYBOOK: {incident['id']}", title_style))
    story.append(Spacer(1, 10))
    
    # Event metadata table
    meta_data = [
        [Paragraph("Incident ID", bold_style), Paragraph(incident['id'], body_style), Paragraph("Corridor", bold_style), Paragraph(incident['corridor'], body_style)],
        [Paragraph("Cause", bold_style), Paragraph(incident['event_cause'].replace('_', ' ').title(), body_style), Paragraph("Junction", bold_style), Paragraph(incident['junction'], body_style)],
        [Paragraph("Priority", bold_style), Paragraph(incident['priority'], body_style), Paragraph("Event Type", bold_style), Paragraph(incident['event_type'].title(), body_style)],
        [Paragraph("Impact Forecast", bold_style), Paragraph(f"{incident['impact_score']} / 10", bold_style), Paragraph("Closure Prob.", bold_style), Paragraph(f"{int(incident['closure_probability'] * 100)}%", body_style)]
    ]
    t = Table(meta_data, colWidths=[110, 150, 110, 150])
    t.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f8fafc')),
        ('BACKGROUND', (2,0), (2,-1), colors.HexColor('#f8fafc')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t)
    story.append(Spacer(1, 15))
    
    # Manpower Optimization
    story.append(Paragraph("1. PU-LP RESOURCE DEPLOYMENT PLAN", h2_style))
    req_off = max(1, int(round(incident['impact_score'] * 1.5)))
    req_bar = int(round(incident['closure_probability'] * 6)) if incident['closure_probability'] >= 0.3 else 0
    
    opt_data = [
        [Paragraph("Resource Class", bold_style), Paragraph("Target Allocation", bold_style), Paragraph("Deployment Directives", bold_style)],
        [Paragraph("Officers", body_style), Paragraph(f"{req_off} Officers", body_style), Paragraph(f"Position at {incident['junction']} for manual signaling.", body_style)],
        [Paragraph("Barricades", body_style), Paragraph(f"{req_bar} Units", body_style), Paragraph("Deploy at entry lanes to execute closures.", body_style)]
    ]
    t_opt = Table(opt_data, colWidths=[120, 120, 280])
    t_opt.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_opt)
    story.append(Spacer(1, 15))
    
    # Route Diversions
    story.append(Paragraph("2. ROUTE DIVERSION RECOMMENDATIONS (NETWORKX SIMULATION)", h2_style))
    div_paths = simulate_flow_diversion((incident['junction'], "HebbalFlyoverJunc" if incident['junction'] != "HebbalFlyoverJunc" else "MekhriCircle"), incident['impact_score'])
    
    div_rows = [
        [Paragraph("Rank", bold_style), Paragraph("Alternative Corridor", bold_style), Paragraph("Added Traffic", bold_style), Paragraph("Est. Travel Time", bold_style)]
    ]
    for div in div_paths[:2]:
        div_rows.append([
            Paragraph(str(div['rank']), body_style),
            Paragraph(', '.join(div['corridors']), body_style),
            Paragraph(f"+{int(div['flow_added'])} units", body_style),
            Paragraph(f"{div['congested_time']:.1f} mins (congested)", body_style)
        ])
        
    t_div = Table(div_rows, colWidths=[40, 180, 140, 160])
    t_div.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_div)
    story.append(Spacer(1, 15))
    
    # Escalation
    story.append(Paragraph("3. ESCALATION & EMERGENCY CONTACTS", h2_style))
    esc_data = [
        [Paragraph("Station/Duty", bold_style), Paragraph("Officer Name", bold_style), Paragraph("Radio Channel", bold_style), Paragraph("Contact", bold_style)],
        [Paragraph("Bengaluru TCC Dispatch", body_style), Paragraph("Inspector Gowda", body_style), Paragraph("VHF Channel 4", body_style), Paragraph("+91 80 2294 2100", body_style)],
        [Paragraph("Field Supervisor", body_style), Paragraph("Sub-Insp. Kumar", body_style), Paragraph("VHF Channel 12", body_style), Paragraph("+91 94808 01000", body_style)]
    ]
    t_esc = Table(esc_data, colWidths=[150, 130, 110, 130])
    t_esc.setStyle(TableStyle([
        ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#f1f5f9')),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('PADDING', (0,0), (-1,-1), 6),
    ]))
    story.append(t_esc)
    
    doc.build(story)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name=f"playbook_{incident['id']}.pdf", mimetype='application/pdf')

@app.route('/')
def index_page():
    return send_file(r'c:\Users\Akshit aggarwal\Downloads\TRAFIQ360\index.html')

@app.route('/inspect_crops')
def inspect_crops():
    return send_file(r'c:\Users\Akshit aggarwal\Downloads\TRAFIQ360\inspect_crops.html')

@app.route('/static/icons/<path:filename>')
def static_icons(filename):
    icons_dir = os.path.join(base_dir, 'static', 'icons')
    return send_file(os.path.join(icons_dir, filename))

# ── FEATURE 1: Post-Event Feedback & Retraining ───────────────────────────────

@app.route('/api/feedback/submit', methods=['POST'])
def feedback_submit():
    data = request.get_json(force=True)
    required = ['incident_id', 'actual_impact', 'actual_duration']
    for f in required:
        if f not in data:
            return jsonify({'success': False, 'error': f'Missing field: {f}'}), 400
    row = {
        'feedback_id': str(uuid.uuid4()),
        'incident_id': data.get('incident_id', ''),
        'event_cause': data.get('event_cause', ''),
        'event_type': data.get('event_type', ''),
        'junction': data.get('junction', ''),
        'corridor': data.get('corridor', ''),
        'predicted_impact': data.get('predicted_impact', ''),
        'actual_impact': data.get('actual_impact', ''),
        'predicted_duration': data.get('predicted_duration', ''),
        'actual_duration': data.get('actual_duration', ''),
        'predicted_resource_count': data.get('predicted_resource_count', ''),
        'actual_resource_count': data.get('actual_resource_count', ''),
        'road_closure_predicted': data.get('road_closure_predicted', ''),
        'road_closure_actual': data.get('road_closure_actual', False),
        'feedback_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'submitted_by': data.get('submitted_by', 'Operator'),
        'approval_status': 'Pending',
        'approved_by': '',
        'approval_timestamp': '',
    }
    _feedback_xlsx_write(row)
    log_audit(data.get('submitted_by', 'Operator'), 'feedback_submit', 'learning',
              {'incident_id': row['incident_id'], 'feedback_id': row['feedback_id']})
    return jsonify({'success': True, 'feedback_id': row['feedback_id'], 'message': 'Feedback submitted and pending approval.'})


@app.route('/api/feedback/list', methods=['GET'])
def feedback_list():
    records = _feedback_xlsx_read()
    # Convert any non-serializable types
    for r in records:
        for k, v in r.items():
            if v is None:
                r[k] = ''
    stats = {
        'total': len(records),
        'pending': sum(1 for r in records if str(r.get('approval_status', '')).lower() == 'pending'),
        'approved': sum(1 for r in records if str(r.get('approval_status', '')).lower() == 'approved'),
        'rejected': sum(1 for r in records if str(r.get('approval_status', '')).lower() == 'rejected'),
    }
    return jsonify({'success': True, 'records': records, 'stats': stats})


@app.route('/api/feedback/approve/<feedback_id>', methods=['POST'])
def feedback_approve(feedback_id):
    data = request.get_json(force=True) or {}
    approver = data.get('approver', 'Admin')
    success = _feedback_xlsx_update(feedback_id, {
        'approval_status': 'Approved',
        'approved_by': approver,
        'approval_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    })
    if success:
        log_audit(approver, 'feedback_approve', 'learning', {'feedback_id': feedback_id})
        return jsonify({'success': True, 'message': 'Feedback approved.'})
    return jsonify({'success': False, 'error': 'Feedback record not found.'}), 404


@app.route('/api/feedback/reject/<feedback_id>', methods=['POST'])
def feedback_reject(feedback_id):
    data = request.get_json(force=True) or {}
    approver = data.get('approver', 'Admin')
    success = _feedback_xlsx_update(feedback_id, {
        'approval_status': 'Rejected',
        'approved_by': approver,
        'approval_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
    })
    if success:
        log_audit(approver, 'feedback_reject', 'learning', {'feedback_id': feedback_id})
        return jsonify({'success': True, 'message': 'Feedback rejected.'})
    return jsonify({'success': False, 'error': 'Feedback record not found.'}), 404


@app.route('/api/feedback/export', methods=['GET'])
def feedback_export_xlsx():
    if not os.path.exists(feedback_xlsx_path):
        return jsonify({'success': False, 'error': 'No feedback data yet.'}), 404
    return send_file(feedback_xlsx_path, as_attachment=True,
                     download_name='trafiq360_feedback.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/feedback/export/csv', methods=['GET'])
def feedback_export_csv():
    records = _feedback_xlsx_read()
    if not records:
        return jsonify({'success': False, 'error': 'No feedback data yet.'}), 404
    buf = io.StringIO()
    writer = csv_module.DictWriter(buf, fieldnames=records[0].keys())
    writer.writeheader()
    writer.writerows(records)
    buf.seek(0)
    return send_file(io.BytesIO(buf.getvalue().encode('utf-8')),
                     as_attachment=True, download_name='trafiq360_feedback.csv',
                     mimetype='text/csv')


@app.route('/api/retrain', methods=['POST'])
def trigger_retrain():
    data = request.get_json(force=True) or {}
    user_role = data.get('user_role', 'Admin')
    job_id = str(uuid.uuid4())[:8]
    retraining_status[job_id] = {'status': 'running', 'progress': 0, 'message': 'Starting...', 'result': None}

    def _progress(pct, msg):
        retraining_status[job_id]['progress'] = pct
        retraining_status[job_id]['message'] = msg

    def _run():
        try:
            sys.path.insert(0, os.path.join(base_dir, 'scratch'))
            from retrain_pipeline import retrain
            result = retrain(progress_callback=_progress)
            retraining_status[job_id]['status'] = 'done'
            retraining_status[job_id]['result'] = result
            load_intelligence_components()
            log_audit(user_role, 'retrain_trigger', 'learning', result)
        except Exception as e:
            retraining_status[job_id]['status'] = 'error'
            retraining_status[job_id]['message'] = str(e)

    t = threading.Thread(target=_run, daemon=True)
    t.start()
    return jsonify({'success': True, 'job_id': job_id, 'message': 'Retraining started in background.'})


@app.route('/api/retrain/status/<job_id>', methods=['GET'])
def retrain_status(job_id):
    if job_id not in retraining_status:
        return jsonify({'success': False, 'error': 'Job not found.'}), 404
    return jsonify({'success': True, **retraining_status[job_id]})


@app.route('/api/models/versions', methods=['GET'])
def model_versions():
    reg = _load_registry()
    return jsonify({'success': True, 'versions': reg.get('versions', [])})


@app.route('/api/training/queue', methods=['GET'])
def training_queue():
    records = _feedback_xlsx_read()
    pending = [r for r in records if str(r.get('approval_status', '')).lower() == 'pending']
    approved = [r for r in records if str(r.get('approval_status', '')).lower() == 'approved']
    rejected = [r for r in records if str(r.get('approval_status', '')).lower() == 'rejected']
    return jsonify({'success': True, 'pending': len(pending), 'approved': len(approved), 'rejected': len(rejected), 'total': len(records)})


# ── FEATURE 2: BTP Events Integration ────────────────────────────────────────

@app.route('/api/btp/events', methods=['GET'])
def btp_events():
    try:
        events = get_btp_events(use_live=True)
        return jsonify({'success': True, 'events': events, 'count': len(events)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/btp/import', methods=['POST'])
def btp_import():
    data = request.get_json(force=True)
    if not data or 'event' not in data:
        return jsonify({'success': False, 'error': 'Missing event data.'}), 400
    event = data['event']
    user_role = data.get('user_role', 'Operator')
    try:
        save_btp_import(event, btp_xlsx_path)
        log_audit(user_role, 'btp_import', 'planner', {'event_id': event.get('event_id', ''), 'event_name': event.get('event_name', '')})
        return jsonify({'success': True, 'message': 'BTP event imported successfully.', 'event': event})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ── FEATURE 4: Audit Log ──────────────────────────────────────────────────────

@app.route('/api/audit/logs', methods=['GET'])
def audit_logs():
    # Return memory + xlsx merged
    from_xlsx = _audit_log_xlsx_read()
    memory_ids = {e['log_id'] for e in audit_log_memory}
    combined = audit_log_memory + [r for r in from_xlsx if r.get('log_id') not in memory_ids]
    combined_sorted = sorted(combined, key=lambda x: str(x.get('timestamp', '')), reverse=True)
    return jsonify({'success': True, 'logs': combined_sorted[:200], 'total': len(combined_sorted)})


@app.route('/api/audit/export', methods=['GET'])
def audit_export():
    if not os.path.exists(audit_xlsx_path):
        return jsonify({'success': False, 'error': 'No audit data yet.'}), 404
    return send_file(audit_xlsx_path, as_attachment=True,
                     download_name='trafiq360_audit.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@app.route('/playbooks/<path:filename>')
def serve_playbooks(filename):
    return send_from_directory(playbooks_dir, filename)

# ── Feature: Production Quality Playbook Generator ────────────────────────────

def capture_map_screenshot(blocked_corridor, blocked_coords, alt_route_coords):
    map_html_path = os.path.join(playbooks_dir, f"temp_map_{uuid.uuid4().hex}.html")
    map_png_path = os.path.join(playbooks_dir, f"map_screenshot_{uuid.uuid4().hex}.png")
    
    try:
        # Determine center
        if blocked_coords and len(blocked_coords) > 0:
            center_lat = sum([c[1] for c in blocked_coords]) / len(blocked_coords)
            center_lon = sum([c[0] for c in blocked_coords]) / len(blocked_coords)
        elif alt_route_coords and len(alt_route_coords) > 0:
            center_lat = sum([c[1] for c in alt_route_coords]) / len(alt_route_coords)
            center_lon = sum([c[0] for c in alt_route_coords]) / len(alt_route_coords)
        else:
            center_lat, center_lon = 12.9716, 77.5946

        m = folium.Map(location=[center_lat, center_lon], zoom_start=14, tiles='cartodbpositron')

        # Add blocked segment
        if blocked_coords:
            # folium expects [lat, lon]
            b_path = [[c[1], c[0]] for c in blocked_coords]
            folium.PolyLine(b_path, color='red', weight=6, dash_array='10, 5', opacity=0.8, popup=f"BLOCKED: {blocked_corridor}").addTo(m)

        # Add alt route
        if alt_route_coords:
            a_path = [[c[1], c[0]] for c in alt_route_coords]
            folium.PolyLine(a_path, color='green', weight=5, opacity=0.8, popup="RECOMMENDED DIVERSION").addTo(m)

        m.save(map_html_path)

        # Selenium headless screenshot
        chrome_options = Options()
        chrome_options.add_argument("--headless")
        chrome_options.add_argument("--window-size=800,600")
        chrome_options.add_argument("--disable-gpu")
        chrome_options.add_argument("--no-sandbox")
        
        driver = webdriver.Chrome(options=chrome_options)
        driver.get(f"file:///{os.path.abspath(map_html_path)}")
        time.sleep(2) # Wait for tiles to load
        driver.save_screenshot(map_png_path)
        driver.quit()
        
        return map_png_path
    except Exception as e:
        print(f"Error capturing map screenshot: {e}")
        return None
    finally:
        if os.path.exists(map_html_path):
            try:
                os.remove(map_html_path)
            except:
                pass

@app.route('/api/generate-playbook', methods=['POST'])
def generate_playbook():
    data = request.json or {}
    pred = data.get('prediction', {})
    opt = data.get('optimization', {})
    div = data.get('diversion', {})
    
    event_cause = pred.get('event_cause', 'Unknown')
    corridor = pred.get('corridor', 'Unknown')
    impact_score = float(pred.get('impact_score', 0))
    closure_prob = float(pred.get('closure_probability', 0))
    expected_duration = float(pred.get('expected_duration', 0))
    priority = pred.get('priority', 'Medium')
    
    timestamp_str = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"playbook_{event_cause}_{corridor.replace(' ', '_')}_{timestamp_str}.pdf"
    filepath = os.path.join(playbooks_dir, filename)

    doc = SimpleDocTemplate(filepath, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=24, spaceAfter=20, alignment=1)
    h2_style = ParagraphStyle('H2Style', parent=styles['Heading2'], fontSize=16, spaceAfter=12, spaceBefore=16, textColor=colors.HexColor('#1f2937'))
    normal_style = styles['Normal']
    
    story = []

    # === PAGE 1: COVER ===
    story.append(Spacer(1, 2*inch))
    story.append(Paragraph("<b>TRAFIQ360 Operational Playbook</b>", title_style))
    story.append(Spacer(1, 0.5*inch))
    
    # Event Details Table
    data_details = [
        ['Cause', event_cause.replace('_', ' ').title()],
        ['Corridor', corridor],
        ['Date', datetime.now().strftime('%Y-%m-%d')],
        ['Hour', str(pred.get('hour', 'N/A'))],
        ['Priority', priority]
    ]
    t = Table(data_details, colWidths=[2*inch, 3*inch])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f3f4f6')),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor('#111827')),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e5e7eb')),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.5*inch))
    
    # Colored impact badge
    if impact_score >= 7:
        bg_color, text_color = '#fee2e2', '#b91c1c'
    elif impact_score >= 5:
        bg_color, text_color = '#fef3c7', '#b45309'
    else:
        bg_color, text_color = '#d1fae5', '#047857'
        
    badge_data = [[f"IMPACT SCORE: {impact_score:.1f} / 10.0"]]
    t_badge = Table(badge_data, colWidths=[5*inch])
    t_badge.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), colors.HexColor(bg_color)),
        ('TEXTCOLOR', (0,0), (-1,-1), colors.HexColor(text_color)),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 16),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('BOX', (0,0), (-1,-1), 1, colors.HexColor(text_color)),
    ]))
    story.append(t_badge)
    
    story.append(Spacer(1, 2*inch))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ParagraphStyle('Footer', parent=normal_style, alignment=1, textColor=colors.gray)))
    story.append(Paragraph("<b>BENGALURU TRAFFIC POLICE — CONFIDENTIAL</b>", ParagraphStyle('FooterBold', parent=normal_style, alignment=1, textColor=colors.gray)))
    
    story.append(PageBreak())

    # === PAGE 2: PREDICTION ===
    story.append(Paragraph("ML Prediction Summary", h2_style))
    story.append(Paragraph(f"<b>Closure Probability:</b> {closure_prob*100:.1f}%", normal_style))
    duration_h = int(expected_duration)
    duration_m = int((expected_duration - duration_h) * 60)
    story.append(Paragraph(f"<b>Expected Duration:</b> {duration_h}h {duration_m}m", normal_style))
    story.append(Spacer(1, 0.2*inch))
    
    story.append(Paragraph("<b>Prediction Drivers (SHAP)</b>", normal_style))
    story.append(Spacer(1, 0.1*inch))
    drivers = pred.get('impact_drivers', [])
    if drivers:
        sh_data = [['Feature', 'Contribution', 'Direction']]
        for d in drivers:
            sh_data.append([str(d.get('feature', '')), f"{d.get('value', 0):.3f}", str(d.get('direction', ''))])
            
        t_shap = Table(sh_data, colWidths=[2.5*inch, 1.5*inch, 1*inch])
        t_shap.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e5e7eb')),
        ]))
        story.append(t_shap)
    else:
        story.append(Paragraph("No driver data available.", normal_style))
        
    story.append(PageBreak())

    # === PAGE 3: RESOURCES ===
    story.append(Paragraph("Resource Deployment Plan", h2_style))
    alloc = opt.get('allocations', [])
    if alloc:
        res_data = [['Junction', 'Officers', 'Barricades', 'Priority']]
        tot_off, tot_bar = 0, 0
        for a in alloc:
            j_name = a.get('junction', 'Unknown')
            off = a.get('officers', 0)
            bar = a.get('barricades', 0)
            tot_off += off
            tot_bar += bar
            res_data.append([j_name, str(off), str(bar), "HIGH" if off > 3 else "MEDIUM"])
            
        t_res = Table(res_data, colWidths=[2.5*inch, 1*inch, 1*inch, 1*inch])
        t_res.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#1d4ed8')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0,0), (-1,-1), 6),
            ('TOPPADDING', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e5e7eb')),
        ]))
        story.append(t_res)
        story.append(Spacer(1, 0.2*inch))
        
        story.append(Paragraph(f"<b>Total Required:</b> {tot_off} Officers | {tot_bar} Barricades", ParagraphStyle('B', parent=normal_style, fontSize=12)))
        story.append(Spacer(1, 0.3*inch))
        
        # Shifts
        story.append(Paragraph("<b>Shift Schedule</b>", normal_style))
        shift_data = [
            ['Shift', 'Start', 'End', 'Officers Required'],
            ['Immediate Response', '0:00 (Now)', '+4:00 hrs', str(tot_off)],
            ['Relief Shift', '+4:00 hrs', '+8:00 hrs', str(tot_off)]
        ]
        t_shift = Table(shift_data, colWidths=[1.5*inch, 1*inch, 1*inch, 1.5*inch])
        t_shift.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4b5563')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e5e7eb')),
        ]))
        story.append(t_shift)
        
    story.append(Spacer(1, 0.4*inch))
    story.append(Paragraph("<b>Escalation Contacts</b>", normal_style))
    story.append(Paragraph("• DCP Traffic East: +91-9876543210", normal_style))
    story.append(Paragraph("• ACP Control Room: +91-8765432109", normal_style))
    
    story.append(PageBreak())

    # === PAGE 4: DIVERSIONS ===
    story.append(Paragraph("Diversion Routes", h2_style))
    
    alt_routes = div.get('alternate_routes', [])
    blocked_seg = div.get('blocked_segment', {})
    
    if alt_routes:
        story.append(Paragraph(f"<b>Blocked Segment:</b> {blocked_seg.get('corridor', 'Unknown')}", normal_style))
        story.append(Spacer(1, 0.2*inch))
        
        div_data = [['Rank', 'Via Corridors', 'Distance', 'Secondary Load', 'Status']]
        for r in alt_routes:
            rank = r.get('rank', '-')
            corrs = ", ".join(r.get('passes_through', []))
            dist = f"{r.get('distance_m', 0)/1000:.1f} km"
            load = f"{r.get('secondary_load_pct', 0):.1f}%"
            rec = r.get('recommendation', '')
            div_data.append([str(rank), corrs, dist, load, rec])
            
        t_div = Table(div_data, colWidths=[0.5*inch, 2*inch, 1*inch, 1.2*inch, 1*inch])
        t_div.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#047857')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('GRID', (0,0), (-1,-1), 1, colors.HexColor('#e5e7eb')),
            ('ALIGN', (2,1), (3,-1), 'RIGHT')
        ]))
        story.append(t_div)
        story.append(Spacer(1, 0.2*inch))
        
        best_route = alt_routes[0]
        story.append(Paragraph(f"<b>Recommended Diversion:</b> Redirect flow to {', '.join(best_route.get('passes_through', []))} ({(best_route.get('distance_m', 0)/1000):.1f} km).", normal_style))
        
        for r in alt_routes:
            if r.get('rank') == 1 and float(r.get('secondary_load_pct', 0)) > 40:
                story.append(Spacer(1, 0.2*inch))
                story.append(Paragraph(f"<b>⚠ WARNING:</b> This diversion overloads {r.get('passes_through', ['the corridor'])[0]} — consider distributing traffic to rank 2 route.", ParagraphStyle('Warn', parent=normal_style, textColor=colors.HexColor('#b45309'))))
    else:
        story.append(Paragraph("No alternate routes calculated.", normal_style))
        
    story.append(PageBreak())

    # === PAGE 5: MAP ===
    story.append(Paragraph("Geospatial Overview", h2_style))
    
    # Attempt to capture map screenshot
    blocked_coords = blocked_seg.get('geojson', {}).get('coordinates', []) if blocked_seg else []
    alt_coords = alt_routes[0].get('geojson', {}).get('coordinates', []) if alt_routes else []
    
    map_png_path = capture_map_screenshot(blocked_seg.get('corridor', 'Unknown'), blocked_coords, alt_coords)
    
    if map_png_path and os.path.exists(map_png_path):
        story.append(RLImage(map_png_path, width=6*inch, height=4.5*inch))
        story.append(Spacer(1, 0.2*inch))
        story.append(Paragraph("Red Dashed: Blocked Incident Corridor | Solid Green: Recommended Diversion Path", ParagraphStyle('Caption', parent=normal_style, alignment=1, fontSize=9, textColor=colors.gray)))
    else:
        story.append(Paragraph("<i>[Map visualization unavailable - Selenium fallback]</i>", normal_style))
        if blocked_coords:
            story.append(Paragraph(f"<b>Blocked segment geometry:</b> {len(blocked_coords)} nodes", normal_style))
            
    doc.build(story)
    
    # Cleanup map image
    if map_png_path and os.path.exists(map_png_path):
        try:
            os.remove(map_png_path)
        except:
            pass

    return jsonify({
        'success': True,
        'download_url': f'/playbooks/{filename}',
        'filename': filename
    })

# ── Feature: Multi-Event Conflict Detection ──────────────────────────────────

def get_zone(corridor):
    c = corridor.lower()
    if 'bellary' in c or 'north' in c: return 'North'
    if 'hosur' in c or 'silk board' in c or 'bommanahalli' in c: return 'South'
    if 'tumkur' in c or 'west' in c or 'yeshwanthpura' in c: return 'West'
    if 'old madras' in c or 'east' in c or 'k r puram' in c: return 'East'
    if 'cbd' in c or 'mysore' in c or 'townhall' in c: return 'Central'
    return 'General'

def get_corridor_nodes(corridor_name):
    nodes = set()
    for u, v, name in CORRIDOR_EDGES:
        if name == corridor_name:
            nodes.add(u)
            nodes.add(v)
    return nodes

@app.route('/api/detect-conflicts', methods=['POST'])
def detect_conflicts():
    data = request.json or {}
    events = data.get('events', [])
    
    conflicts = []
    severity_summary = {"CRITICAL": 0, "HIGH": 0, "MEDIUM": 0}
    joint_opt_recommended = False
    
    # 1. Global RESOURCE_CONFLICT check
    total_officers_needed = 0
    available_officers = 20 # default
    
    if events:
        available_officers = int(events[0].get('available_officers', 20))
        
    for ev in events:
        impact = float(ev.get('impact_score', 5.0))
        total_officers_needed += int(max(1, round(impact * 1.5)))
        
    if total_officers_needed > available_officers:
        conflicts.append({
            "type": "RESOURCE_CONFLICT",
            "severity": "CRITICAL",
            "events": [ev.get('id', 'Unknown') for ev in events],
            "message": f"Combined officer requirement ({total_officers_needed}) exceeds availability ({available_officers})",
            "fix": "Request mutual aid from adjacent zones"
        })
        severity_summary["CRITICAL"] += 1
        joint_opt_recommended = True

    # Check pairwise conflicts
    n = len(events)
    for i in range(n):
        for j in range(i + 1, n):
            ev1, ev2 = events[i], events[j]
            c1, c2 = ev1.get('corridor', ''), ev2.get('corridor', '')
            h1, h2 = int(ev1.get('hour', 0)), int(ev2.get('hour', 0))
            imp1, imp2 = float(ev1.get('impact_score', 0)), float(ev2.get('impact_score', 0))
            z1, z2 = ev1.get('zone') or get_zone(c1), ev2.get('zone') or get_zone(c2)
            
            # CORRIDOR_CONFLICT
            if c1 == c2 and abs(h1 - h2) <= 4:
                sev = "HIGH" if imp1 >= 6 and imp2 >= 6 else "MEDIUM"
                conflicts.append({
                    "type": "CORRIDOR_CONFLICT",
                    "severity": sev,
                    "events": [ev1.get('id'), ev2.get('id')],
                    "message": f"Two events on {c1} within {abs(h1 - h2)}h — officer pool conflict",
                    "fix": "Run joint ILP optimization with combined officer pool"
                })
                severity_summary[sev] += 1
                joint_opt_recommended = True
            
            # CASCADE_RISK
            if c1 != c2:
                nodes1 = get_corridor_nodes(c1)
                nodes2 = get_corridor_nodes(c2)
                if nodes1.intersection(nodes2):
                    conflicts.append({
                        "type": "CASCADE_RISK",
                        "severity": "MEDIUM",
                        "events": [ev1.get('id'), ev2.get('id')],
                        "message": f"Adjacent corridors blocked ({c1} and {c2}) — cascade congestion likely",
                        "fix": "Activate secondary diversion loops immediately"
                    })
                    severity_summary["MEDIUM"] += 1
                    
            # ZONE_CONFLICT
            if z1 == z2 and z1 != 'General':
                # only if they overlap somewhat in time, let's say within 6 hours
                if abs(h1 - h2) <= 6:
                    conflicts.append({
                        "type": "ZONE_CONFLICT",
                        "severity": "MEDIUM",
                        "events": [ev1.get('id'), ev2.get('id')],
                        "message": f"Multiple events in zone {z1} — DCP coordination required",
                        "fix": "Alert Zone Commander"
                    })
                    severity_summary["MEDIUM"] += 1

    return jsonify({
        "success": True,
        "conflicts": conflicts,
        "severity_summary": severity_summary,
        "joint_optimization_recommended": joint_opt_recommended
    })

# Startup
if __name__ == '__main__':
    if not os.path.exists(playbooks_dir):
        os.makedirs(playbooks_dir)
    app.run(debug=True, port=5000)
