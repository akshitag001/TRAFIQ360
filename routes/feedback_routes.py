from flask import Blueprint, request, jsonify, send_file
import os
import io
import uuid
import csv as csv_module
from datetime import datetime
from core.config import feedback_xlsx_path
from routes.audit_routes import log_audit

feedback_bp = Blueprint('feedback', __name__)

def _feedback_xlsx_write(row_data: dict):
    """Append a feedback row to the Excel file."""
    import openpyxl
    COLUMNS = [
        'feedback_id','incident_id','event_cause','event_type','junction','corridor',
        'predicted_impact','actual_impact','predicted_duration','actual_duration',
        'predicted_resource_count','actual_resource_count',
        'road_closure_predicted','road_closure_actual',
        'feedback_timestamp','submitted_by','approval_status','approved_by','approval_timestamp'
    ]
    if os.path.exists(feedback_xlsx_path):
        wb = openpyxl.load_workbook(feedback_xlsx_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Feedback"
        ws.append(COLUMNS)
    ws.append([row_data.get(c, '') for c in COLUMNS])
    wb.save(feedback_xlsx_path)

def _feedback_xlsx_read() -> list:
    import openpyxl
    if not os.path.exists(feedback_xlsx_path):
        return []
    wb = openpyxl.load_workbook(feedback_xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]

def _feedback_xlsx_update(feedback_id: str, update: dict):
    import openpyxl
    if not os.path.exists(feedback_xlsx_path):
        return False
    wb = openpyxl.load_workbook(feedback_xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    id_col = headers.index('feedback_id') + 1
    for row in ws.iter_rows(min_row=2):
        if str(row[id_col - 1].value) == feedback_id:
            for key, val in update.items():
                if key in headers:
                    row[headers.index(key)].value = val
            break
    wb.save(feedback_xlsx_path)
    return True

@feedback_bp.route('/api/feedback/submit', methods=['POST'])
def feedback_submit():
    try:
        data = request.get_json(force=True)
        required = ['incident_id', 'actual_impact', 'actual_duration']
        for f in required:
            if f not in data:
                return jsonify({'success': False, 'error': f'Missing field: {f}'}), 400
        row = {
            'feedback_id': str(uuid.uuid4()),
            'incident_id': data.get('incident_id', ''),
            'event_cause': data.get('event_cause', ''),
            'event_type': data.get('event_type', ''),
            'junction': data.get('junction', ''),
            'corridor': data.get('corridor', ''),
            'predicted_impact': data.get('predicted_impact', ''),
            'actual_impact': data.get('actual_impact', ''),
            'predicted_duration': data.get('predicted_duration', ''),
            'actual_duration': data.get('actual_duration', ''),
            'predicted_resource_count': data.get('predicted_resource_count', ''),
            'actual_resource_count': data.get('actual_resource_count', ''),
            'road_closure_predicted': data.get('road_closure_predicted', ''),
            'road_closure_actual': data.get('road_closure_actual', False),
            'feedback_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'submitted_by': data.get('submitted_by', 'Operator'),
            'approval_status': 'Pending',
            'approved_by': '',
            'approval_timestamp': '',
        }
        _feedback_xlsx_write(row)
        log_audit(data.get('submitted_by', 'Operator'), 'feedback_submit', 'learning',
                  {'incident_id': row['incident_id'], 'feedback_id': row['feedback_id']})
        return jsonify({'success': True, 'feedback_id': row['feedback_id'], 'message': 'Feedback submitted and pending approval.'})
    except Exception as e:
        print(f"[ERROR] /api/feedback/submit failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@feedback_bp.route('/api/feedback/list', methods=['GET'])
def feedback_list():
    try:
        records = _feedback_xlsx_read()
        for r in records:
            for k, v in r.items():
                if v is None:
                    r[k] = ''
        stats = {
            'total': len(records),
            'pending': sum(1 for r in records if str(r.get('approval_status', '')).lower() == 'pending'),
            'approved': sum(1 for r in records if str(r.get('approval_status', '')).lower() == 'approved'),
            'rejected': sum(1 for r in records if str(r.get('approval_status', '')).lower() == 'rejected'),
        }
        return jsonify({'success': True, 'records': records, 'stats': stats})
    except Exception as e:
        print(f"[ERROR] /api/feedback/list failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@feedback_bp.route('/api/feedback/approve/<feedback_id>', methods=['POST'])
def feedback_approve(feedback_id):
    try:
        data = request.get_json(force=True) or {}
        approver = data.get('approver', 'Admin')
        success = _feedback_xlsx_update(feedback_id, {
            'approval_status': 'Approved',
            'approved_by': approver,
            'approval_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        if success:
            log_audit(approver, 'feedback_approve', 'learning', {'feedback_id': feedback_id})
            return jsonify({'success': True, 'message': 'Feedback approved.'})
        return jsonify({'success': False, 'error': 'Feedback record not found.'}), 404
    except Exception as e:
        print(f"[ERROR] /api/feedback/approve failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@feedback_bp.route('/api/feedback/reject/<feedback_id>', methods=['POST'])
def feedback_reject(feedback_id):
    try:
        data = request.get_json(force=True) or {}
        approver = data.get('approver', 'Admin')
        success = _feedback_xlsx_update(feedback_id, {
            'approval_status': 'Rejected',
            'approved_by': approver,
            'approval_timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        })
        if success:
            log_audit(approver, 'feedback_reject', 'learning', {'feedback_id': feedback_id})
            return jsonify({'success': True, 'message': 'Feedback rejected.'})
        return jsonify({'success': False, 'error': 'Feedback record not found.'}), 404
    except Exception as e:
        print(f"[ERROR] /api/feedback/reject failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@feedback_bp.route('/api/feedback/export', methods=['GET'])
def feedback_export_xlsx():
    try:
        if not os.path.exists(feedback_xlsx_path):
            return jsonify({'success': False, 'error': 'No feedback data yet.'}), 404
        return send_file(feedback_xlsx_path, as_attachment=True,
                         download_name='trafiq360_feedback.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        print(f"[ERROR] /api/feedback/export failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@feedback_bp.route('/api/feedback/export/csv', methods=['GET'])
def feedback_export_csv():
    try:
        records = _feedback_xlsx_read()
        if not records:
            return jsonify({'success': False, 'error': 'No feedback data yet.'}), 404
        buf = io.StringIO()
        writer = csv_module.DictWriter(buf, fieldnames=records[0].keys())
        writer.writeheader()
        writer.writerows(records)
        buf.seek(0)
        return send_file(io.BytesIO(buf.getvalue().encode('utf-8')),
                         as_attachment=True, download_name='trafiq360_feedback.csv',
                         mimetype='text/csv')
    except Exception as e:
        print(f"[ERROR] /api/feedback/export/csv failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
