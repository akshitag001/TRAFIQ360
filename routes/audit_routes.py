from flask import Blueprint, jsonify
import os
import uuid
from datetime import datetime
import json
from core.config import audit_xlsx_path
from core.model_loader import get_state

audit_bp = Blueprint('audit', __name__)

def _audit_log_xlsx_write(entry: dict):
    import openpyxl
    COLS = ['log_id','timestamp','user_role','action','module','details']
    if os.path.exists(audit_xlsx_path):
        wb = openpyxl.load_workbook(audit_xlsx_path)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "AuditLog"
        ws.append(COLS)
    ws.append([entry.get(c, '') for c in COLS])
    wb.save(audit_xlsx_path)

def _audit_log_xlsx_read() -> list:
    import openpyxl
    if not os.path.exists(audit_xlsx_path):
        return []
    wb = openpyxl.load_workbook(audit_xlsx_path)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h) for h in rows[0]]
    return [dict(zip(headers, row)) for row in rows[1:]]

def log_audit(user_role: str, action: str, module: str, details: dict = None):
    """Write audit entry to memory + xlsx."""
    entry = {
        'log_id': str(uuid.uuid4())[:8],
        'timestamp': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'user_role': user_role,
        'action': action,
        'module': module,
        'details': json.dumps(details or {})
    }
    state = get_state()
    state['audit_log_memory'].append(entry)
    try:
        _audit_log_xlsx_write(entry)
    except Exception:
        pass

@audit_bp.route('/api/audit/logs', methods=['GET'])
def get_audit_logs():
    try:
        # Priority to memory, fallback to disk
        state = get_state()
        logs = state['audit_log_memory']
        if not logs:
            logs = _audit_log_xlsx_read()
            state['audit_log_memory'] = logs  # populate memory
        return jsonify({'success': True, 'logs': logs})
    except Exception as e:
        print(f"[ERROR] /api/audit/logs failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500

@audit_bp.route('/api/audit/export', methods=['GET'])
def export_audit_logs():
    try:
        from flask import send_file
        if not os.path.exists(audit_xlsx_path):
            return jsonify({'success': False, 'message': 'No audit logs found to export'}), 404
        return send_file(audit_xlsx_path, as_attachment=True, download_name="TRAFIQ360_Audit_Logs.xlsx")
    except Exception as e:
        print(f"[ERROR] /api/audit/export failed: {str(e)}")
        return jsonify({'success': False, 'error': str(e)}), 500
